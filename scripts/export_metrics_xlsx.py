"""Consolidate metrics/data CSVs into reports/wheel_metrics.xlsx.

Sheets:
  Summary         — headline numbers with cross-sheet formulas
  DailySnapshot   — real account equity / cash / BP / daily P&L
  ShadowDecisions — daily v3 rotation decisions + per-symbol scores
  ShadowEquity    — virtual account equity over time
  ShadowRotations — every virtual rotation event
"""
from __future__ import annotations
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
DATA = ROOT / "metrics" / "data"
OUT = ROOT / "reports" / "wheel_metrics.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F3864")
CELL_FONT = Font(name="Arial", size=10)
BORDER = Border(bottom=Side(style="thin", color="BFBFBF"))

CURRENCY_FMT = '$#,##0.00;($#,##0.00);-'
PCT_FMT = '0.00%;(0.00%);-'
DATE_FMT = 'yyyy-mm-dd'


def _load_csv(name: str) -> pd.DataFrame | None:
    path = DATA / name
    if not path.exists():
        return None
    df = pd.read_csv(path)
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
    if "closed_at" in df.columns:
        df["closed_at"] = pd.to_datetime(df["closed_at"], errors="coerce")
    return df


def _write_sheet(wb: Workbook, name: str, df: pd.DataFrame,
                  currency_cols: set[str], pct_cols: set[str],
                  date_cols: set[str]) -> None:
    ws = wb.create_sheet(name)
    headers = list(df.columns)
    ws.append(headers)

    # Header styling
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row in df.itertuples(index=False, name=None):
        ws.append(row)

    # Column formats
    for col_idx, hdr in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        if hdr in currency_cols:
            fmt = CURRENCY_FMT
        elif hdr in pct_cols:
            fmt = PCT_FMT
        elif hdr in date_cols:
            fmt = DATE_FMT
        else:
            fmt = None

        max_len = len(str(hdr))
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = CELL_FONT
            cell.border = BORDER
            if fmt:
                cell.number_format = fmt
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)[:25]))
        ws.column_dimensions[letter].width = min(max_len + 2, 30)

    ws.freeze_panes = "A2"


def _build_summary(wb: Workbook, sheet_row_counts: dict[str, int]) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Wheel Metrics — Consolidated Report"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws.merge_cells("A1:C1")

    # Use COUNTA-based last-row formulas so the Summary auto-follows data growth
    # even if the sheets are appended to manually (not just regenerated).
    #   real equity col = C, shadow equity col = B
    #   real date col   = A, shadow date col   = A

    rows = [
        ("", "", ""),
        ("Real account", "", ""),
        ("Current equity",
         "=IFERROR(INDEX(DailySnapshot!C:C,COUNTA(DailySnapshot!A:A)),\"n/a\")",
         "currency"),
        ("Most recent snapshot",
         "=IFERROR(INDEX(DailySnapshot!A:A,COUNTA(DailySnapshot!A:A)),\"n/a\")",
         "date"),
        ("Days tracked",
         "=MAX(0,COUNTA(DailySnapshot!A:A)-1)", "number"),
        ("", "", ""),
        ("Shadow account (v3 virtual)", "", ""),
        ("Current equity",
         "=IFERROR(INDEX(ShadowEquity!B:B,COUNTA(ShadowEquity!A:A)),\"n/a\")",
         "currency"),
        ("Most recent snapshot",
         "=IFERROR(INDEX(ShadowEquity!A:A,COUNTA(ShadowEquity!A:A)),\"n/a\")",
         "date"),
        ("Days tracked",
         "=MAX(0,COUNTA(ShadowEquity!A:A)-1)", "number"),
        ("Total rotations recorded",
         "=MAX(0,COUNTA(ShadowRotations!A:A)-1)", "number"),
        ("", "", ""),
        ("Delta (Shadow - Real)", "", ""),
        ("Equity difference ($)",
         "=IFERROR(B9-B4,\"n/a\")", "currency"),
        ("Equity difference (%)",
         "=IFERROR((B9-B4)/B4,\"n/a\")", "pct"),
    ]

    for i, (label, val, kind) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = (
            Font(name="Arial", bold=True, size=11)
            if not label.startswith(" ") and label and val == ""
            else Font(name="Arial", size=10)
        )
        if val == "":
            continue
        c = ws.cell(row=i, column=2, value=val)
        c.font = CELL_FONT
        if kind == "currency":
            c.number_format = CURRENCY_FMT
        elif kind == "pct":
            c.number_format = PCT_FMT
        elif kind == "date":
            c.number_format = DATE_FMT

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12


def main() -> int:
    sheets_config = {
        "DailySnapshot": {
            "csv": "daily_snapshot.csv",
            "currency": {"equity", "last_equity", "cash", "buying_power", "daily_pl"},
            "pct": {"daily_pl_pct"},
            "date": {"date"},
        },
        "ShadowDecisions": {
            "csv": "shadow_decisions.csv",
            "currency": {"equity", "cash", "premium_total"},
            "pct": set(),
            "date": {"date"},
        },
        "ShadowEquity": {
            "csv": "shadow_equity.csv",
            "currency": {"shadow_equity", "cash", "spot_price"},
            "pct": set(),
            "date": {"date"},
        },
        "ShadowRotations": {
            "csv": "shadow_rotations.csv",
            "currency": set(),
            "pct": set(),    # improvement column is already a string "25.2%"
            "date": {"date"},
        },
    }

    wb = Workbook()
    # Remove default empty sheet — Summary will be prepended later
    default = wb.active
    wb.remove(default)

    row_counts: dict[str, int] = {}
    for name, cfg in sheets_config.items():
        df = _load_csv(cfg["csv"])
        if df is None or df.empty:
            print(f"  skip {name} (no data)")
            row_counts[name] = 0
            continue
        _write_sheet(wb, name, df, cfg["currency"], cfg["pct"], cfg["date"])
        row_counts[name] = len(df)
        print(f"  {name}: {len(df)} rows")

    _build_summary(wb, row_counts)

    wb.save(OUT)
    print(f"\n✓ Wrote {OUT.relative_to(ROOT)} ({OUT.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
